from datetime import datetime as dt
import pandas as pd
import pytz
import openpyxl
from astral import moon
import requests

BASE_URL = 'http://api.openweathermap.org/data/2.5/weather?'
api_key = '********************************'

####################################
#####- Date/time Calculations -#####
####################################
current_date = dt.now()
current_year = current_date.strftime('%Y')
current_month = current_date.strftime('%m')
current_day = current_date.strftime('%d')
current_time_CST = current_date.strftime('%H%M')
current_time_MDT = dt.now(pytz.timezone('America/Denver'))
current_time_MDT = current_time_MDT.strftime('%H%M')

#####################################
#####- Lunar phase Calculation -#####
#####################################
lunar_phase_data = moon.phase(dt.now())
if 0 <= lunar_phase_data <= 3.49:
    lunar_phase = 'New Moon'
elif 3.49 <= lunar_phase_data <= 6.99:
    lunar_phase = 'Waxing Crescent'
elif 7 <= lunar_phase_data <= 10.49:
    lunar_phase = 'First Quarter'
elif 10.5 <= lunar_phase_data <= 13.99:
    lunar_phase = 'Waxing Gibbous'
elif 14.00 <= lunar_phase_data <= 17.49:
    lunar_phase = 'Full Moon'
elif 17.5 <= lunar_phase_data <= 20.99:
    lunar_phase = 'Waning Gibbous'
elif 21.00 <= lunar_phase_data <= 24.49:
    lunar_phase = 'Last Quarter'
elif 24.5 <= lunar_phase_data <= 27.99:
    lunar_phase = 'Waning Crescent'
else:
    print('Invalid lunar phase.')
    print(lunar_phase_data)
    lunar_phase = '<Null>'

########################################
#####- Wind direction calculation -#####
########################################

def convert_wind_direction(wind_direction_data):
    wind_direction_deg = round(wind_direction_data, 0)
    if 337.6 <= wind_direction_deg <= 360:
        return 'N'
    elif 0 <= wind_direction_deg <= 22.5:
        return 'N'
    elif 22.6 <= wind_direction_deg <= 67.5:
        return 'NE'
    elif 67.6 <= wind_direction_deg <= 112.5:
        return 'E'
    elif 112.6 <= wind_direction_deg <= 157.5:
        return 'SE'
    elif 157.6 <= wind_direction_deg <= 202.5:
        return 'S'
    elif 202.6 <= wind_direction_deg <= 247.5:
        return 'SW'
    elif 247.6 <= wind_direction_deg <= 292.5:
        return 'W'
    elif 292.6 <= wind_direction_deg <= 337.5:
        return 'NW'
    else:
        print('Invalid wind direction.')
        return '<Null>'

def get_weather_for_city(city, blank_string):
    url = f'{BASE_URL}appid={api_key}&q={city}'
    response = requests.get(url).json()
    try:
        temp_fahrenheit = round((((response['main']['temp']) - 273.15) * (9/5) + 32), 2) #- Converts from temp K to temp F
    except:
        temp_fahrenheit = "<Null>"
    try:
        feels_like_temp_fahrenheit = round((((response['main']['feels_like']) - 273.15) * (9/5) + 32), 2) #- Converts from temp K to temp F
    except:
        feels_like_temp_fahrenheit = blank_string
    try:
        low_temp_fahrenheit = round((((response['main']['temp_min']) - 273.15) * (9/5) + 32), 2) #- Converts from temp K to temp F
    except:
        low_temp_fahrenheit = blank_string  
    try:
        high_temp_fahrenheit = round((((response['main']['temp_max']) - 273.15) * (9/5) + 32), 2) #- Converts from temp K to temp F
    except:
        high_temp_fahrenheit = blank_string      
    try:
        wind_speed = response['wind']['speed']
    except:
        wind_speed = blank_string
    try:
        wind_direction = response['wind']['deg']
    except:
        wind_direction = blank_string
    try:
        wind_gust = response['wind']['gust']
    except:
        wind_gust = blank_string
    try:
        pressure = response['main']['pressure']
    except:
        pressure = blank_string
    try:
        description = (response['weather'][0]['description']).title()
    except:
        description = "<Null>"
    try:
        humidity = response['main']['humidity']
    except:
        humidity = blank_string
    try:
        sunrise_timestamp = response['sys']['sunrise']
        sunrise_read_dt_object = dt.fromtimestamp(sunrise_timestamp)
        sunrise_time = sunrise_read_dt_object.strftime('%H:%M')
    except:
        sunrise_time = blank_string
    try:
        sunset_timestamp = response['sys']['sunset']
        sunset_read_dt_object = dt.fromtimestamp(sunset_timestamp)
        sunset_time = sunset_read_dt_object.strftime('%H:%M')
    except:
        sunset_time = blank_string
    try:
        start = dt.strptime(sunrise_time, '%H:%M')
        end = dt.strptime(sunset_time, '%H:%M')
        hours_of_sunlight = end - start
    except:
        hours_of_sunlight = blank_string
    try:
        visibility = round((response['visibility'] * 0.000621371), 2)
    except:
        visibility = blank_string
    try:
        latitude = response['coord']['lat']
    except:
        latitude = blank_string
    try:
        longitude = response['coord']['lon']
    except:
        longitude = blank_string
    return(city, temp_fahrenheit, feels_like_temp_fahrenheit, low_temp_fahrenheit, high_temp_fahrenheit, wind_speed, wind_gust, wind_direction, pressure, humidity, description, sunrise_time, sunset_time, hours_of_sunlight, visibility, latitude, longitude)

#############################
#####- DF Construction -#####
#############################
# - Rockport, Texas
city, temp_fahrenheit, feels_like_temp_fahrenheit, low_temp_fahrenheit, high_temp_fahrenheit, wind_speed, wind_gust, wind_direction, pressure, humidity, description, sunrise_time, sunset_time, hours_of_sunlight, visibility, latitude, longitude = get_weather_for_city('Rockport', '<Null>')
current_RP_empty_list = []
current_RP_dictionary = {'City': 'Rockport, Texas',
                            'Latitude': latitude,
                            'Longitude': longitude,
                            'Year': current_year,
                            'Month': current_month,
                            'Date': current_day,
                            'Time': current_time_CST,
                            'Description': description,
                            'Temperature': temp_fahrenheit,
                            'Temperature (feels like)': feels_like_temp_fahrenheit,
                            'Temp Low': low_temp_fahrenheit,
                            'Temp High': high_temp_fahrenheit,
                            'Wind Speed (mph)': wind_speed,
                            'Gust Speed': wind_gust,
                            'Wind Direction (deg)': wind_direction,
                            'Wind Direction': convert_wind_direction(wind_direction),
                            'Pressure (mb)': pressure,
                            'Humidity': humidity,
                            'Sunrise': sunrise_time,
                            'Sunset': sunset_time,
                            'Hours of Daylight': hours_of_sunlight,
                            'Visibility': visibility,
                            'Lunar Phase Age': round(lunar_phase_data, 2),
                            'Lunar Phase': lunar_phase
                            }
current_RP_empty_list.append(current_RP_dictionary)
current_RP_df = pd.DataFrame(current_RP_empty_list)

# - Wichita Falls, Texas
city, temp_fahrenheit, feels_like_temp_fahrenheit, low_temp_fahrenheit, high_temp_fahrenheit, wind_speed, wind_gust, wind_direction, pressure, humidity, description, sunrise_time, sunset_time, hours_of_sunlight, visibility, latitude, longitude = get_weather_for_city('Wichita Falls', '<Null>')
current_WF_empty_list = []
current_WF_dictionary = {'City': 'Wichita Falls, Texas',
                            'Latitude': latitude,
                            'Longitude': longitude,
                            'Year': current_year,
                            'Month': current_month,
                            'Date': current_day,
                            'Time': current_time_CST,
                            'Description': description,
                            'Temperature': temp_fahrenheit,
                            'Temperature (feels like)': feels_like_temp_fahrenheit,
                            'Temp Low': low_temp_fahrenheit,
                            'Temp High': high_temp_fahrenheit,
                            'Wind Speed (mph)': wind_speed,
                            'Gust Speed': wind_gust,
                            'Wind Direction (deg)': wind_direction,
                            'Wind Direction': convert_wind_direction(wind_direction),
                            'Pressure (mb)': pressure,
                            'Humidity': humidity,
                            'Sunrise': sunrise_time,
                            'Sunset': sunset_time,
                            'Hours of Daylight': hours_of_sunlight,
                            'Visibility': visibility,
                            'Lunar Phase Age': round(lunar_phase_data, 2),
                            'Lunar Phase': lunar_phase
                            }
current_WF_empty_list.append(current_WF_dictionary)
current_WF_df = pd.DataFrame(current_WF_empty_list)

# - Dodge City, Kansas
city, temp_fahrenheit, feels_like_temp_fahrenheit, low_temp_fahrenheit, high_temp_fahrenheit, wind_speed, wind_gust, wind_direction, pressure, humidity, description, sunrise_time, sunset_time, hours_of_sunlight, visibility, latitude, longitude = get_weather_for_city('Dodge City', '<Null>')
current_DC_empty_list = []
current_DC_dictionary = {'City': 'Dodge City, KS',
                            'Latitude': latitude,
                            'Longitude': longitude,
                            'Year': current_year,
                            'Month': current_month,
                            'Date': current_day,
                            'Time': current_time_CST,
                            'Description': description,
                            'Temperature': temp_fahrenheit,
                            'Temperature (feels like)': feels_like_temp_fahrenheit,
                            'Temp Low': low_temp_fahrenheit,
                            'Temp High': high_temp_fahrenheit,
                            'Wind Speed (mph)': wind_speed,
                            'Gust Speed': wind_gust,
                            'Wind Direction (deg)': wind_direction,
                            'Wind Direction': convert_wind_direction(wind_direction),
                            'Pressure (mb)': pressure,
                            'Humidity': humidity,
                            'Sunrise': sunrise_time,
                            'Sunset': sunset_time,
                            'Hours of Daylight': hours_of_sunlight,
                            'Visibility': visibility,
                            'Lunar Phase Age': round(lunar_phase_data, 2),
                            'Lunar Phase': lunar_phase
                            }
current_DC_empty_list.append(current_DC_dictionary)
current_DC_df = pd.DataFrame(current_DC_empty_list)

# - Scottsbluff, NE
city, temp_fahrenheit, feels_like_temp_fahrenheit, low_temp_fahrenheit, high_temp_fahrenheit, wind_speed, wind_gust, wind_direction, pressure, humidity, description, sunrise_time, sunset_time, hours_of_sunlight, visibility, latitude, longitude = get_weather_for_city('Scottsbluff', '<Null>')
current_SB_empty_list = []
current_SB_dictionary = {'City': 'Scottsbluff, NE',
                            'Latitude': latitude,
                            'Longitude': longitude,
                            'Year': current_year,
                            'Month': current_month,
                            'Date': current_day,
                            'Time': current_time_MDT,
                            'Description': description,
                            'Temperature': temp_fahrenheit,
                            'Temperature (feels like)': feels_like_temp_fahrenheit,
                            'Temp Low': low_temp_fahrenheit,
                            'Temp High': high_temp_fahrenheit,
                            'Wind Speed (mph)': wind_speed,
                            'Gust Speed': wind_gust,
                            'Wind Direction (deg)': wind_direction,
                            'Wind Direction': convert_wind_direction(wind_direction),
                            'Pressure (mb)': pressure,
                            'Humidity': humidity,
                            'Sunrise': sunrise_time,
                            'Sunset': sunset_time,
                            'Hours of Daylight': hours_of_sunlight,
                            'Visibility': visibility,
                            'Lunar Phase Age': round(lunar_phase_data, 2),
                            'Lunar Phase': lunar_phase
                            }
current_SB_empty_list.append(current_SB_dictionary)
current_SB_df = pd.DataFrame(current_SB_empty_list)

# - Glendive, MT
city, temp_fahrenheit, feels_like_temp_fahrenheit, low_temp_fahrenheit, high_temp_fahrenheit, wind_speed, wind_gust, wind_direction, pressure, humidity, description, sunrise_time, sunset_time, hours_of_sunlight, visibility, latitude, longitude = get_weather_for_city('Glendive', '<Null>')
current_GD_empty_list = []
current_GD_dictionary = {'City': 'Glendive, MT',
                            'Latitude': latitude,
                            'Longitude': longitude,
                            'Year': current_year,
                            'Month': current_month,
                            'Date': current_day,
                            'Time': current_time_MDT,
                            'Description': description,
                            'Temperature': temp_fahrenheit,
                            'Temperature (feels like)': feels_like_temp_fahrenheit,
                            'Temp Low': low_temp_fahrenheit,
                            'Temp High': high_temp_fahrenheit,
                            'Wind Speed (mph)': wind_speed,
                            'Gust Speed': wind_gust,
                            'Wind Direction (deg)': wind_direction,
                            'Wind Direction': convert_wind_direction(wind_direction),
                            'Pressure (mb)': pressure,
                            'Humidity': humidity,
                            'Sunrise': sunrise_time,
                            'Sunset': sunset_time,
                            'Hours of Daylight': hours_of_sunlight,
                            'Visibility': visibility,
                            'Lunar Phase Age': round(lunar_phase_data, 2),
                            'Lunar Phase': lunar_phase
                            }
current_GD_empty_list.append(current_GD_dictionary)
current_GD_df = pd.DataFrame(current_GD_empty_list)

# - Saskatoon, SK
city, temp_fahrenheit, feels_like_temp_fahrenheit, low_temp_fahrenheit, high_temp_fahrenheit, wind_speed, wind_gust, wind_direction, pressure, humidity, description, sunrise_time, sunset_time, hours_of_sunlight, visibility, latitude, longitude = get_weather_for_city('Saskatoon', '<Null>')
current_SK_empty_list = []
current_SK_dictionary = {'City': 'Saskatoon, SK',
                            'Latitude': latitude,
                            'Longitude': longitude,
                            'Year': current_year,
                            'Month': current_month,
                            'Date': current_day,
                            'Time': current_time_CST,
                            'Description': description,
                            'Temperature': temp_fahrenheit,
                            'Temperature (feels like)': feels_like_temp_fahrenheit,
                            'Temp Low': low_temp_fahrenheit,
                            'Temp High': high_temp_fahrenheit,
                            'Wind Speed (mph)': wind_speed,
                            'Gust Speed': wind_gust,
                            'Wind Direction (deg)': wind_direction,
                            'Wind Direction': convert_wind_direction(wind_direction),
                            'Pressure (mb)': pressure,
                            'Humidity': humidity,
                            'Sunrise': sunrise_time,
                            'Sunset': sunset_time,
                            'Hours of Daylight': hours_of_sunlight,
                            'Visibility': visibility,
                            'Lunar Phase Age': round(lunar_phase_data, 2),
                            'Lunar Phase': lunar_phase
                            }
current_SK_empty_list.append(current_SK_dictionary)
current_SK_df = pd.DataFrame(current_SK_empty_list)

######################################
#####- Updating the spreadsheet -#####
######################################
excel_path = r'C:\Path\And\File\Name\Of\Excel.xlsx'
wb = openpyxl.load_workbook(excel_path)
RP_Sheet = wb['Rockport, TX']
WF_Sheet = wb['Wichita Falls, TX']
DC_Sheet = wb['Dodge City, KS']
SB_Sheet = wb['Scottsbluff, NE']
GD_Sheet = wb['Glendive, MT']
SK_Sheet = wb['Saskatoon, SK']

def update_spreadsheet(dataframe, worksheet):
    # Find the next empty row in the sheet
    next_row = worksheet.max_row + 1
    for r_idx, row in enumerate(dataframe.itertuples(index=False, name=None), start=next_row):
        for c_idx, value in enumerate(row, start=1):
            worksheet.cell(row=r_idx, column=c_idx, value=value)

update_spreadsheet(current_RP_df, RP_Sheet)
update_spreadsheet(current_WF_df, WF_Sheet)
update_spreadsheet(current_DC_df, DC_Sheet)
update_spreadsheet(current_SB_df, SB_Sheet)
update_spreadsheet(current_GD_df, GD_Sheet)
update_spreadsheet(current_SK_df, SK_Sheet)

# Save the workbook
wb.save(excel_path)
print('Data appended and file saved.')
print(f'Ran at: {current_time_CST}\n')
